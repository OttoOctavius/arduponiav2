# -*- coding: utf-8 -*-
import openpyxl
import serial

#Se uso el link https://automatetheboringstuff.com/chapter12/ de referencia
#ser = serial.Serial('/dev/ttyACM0',9600, timeout=1)

def ciclograbado(archivo,puerto):
    ser = serial.Serial('/dev/ttyACM1') 
    grabar(archivo,ser)
    ser.close()

def grabar(archivo,ser):  
    with open(archivo + ".txt","w") as file:
        while(True):
            if ser.inWaiting():
                try:
                    line = ser.readline().decode('utf-8')
                    file.write( line)
                except:
                    Warning('No se pudo leer')
                    break
            


def crearArchivo():
    wb = openpyxl.Workbook()
    wb.save('Menta.xlsx')

#diccionario = {'a':[1,2,3],'b':[1,2,3]} , clave:[lista]
def salvarEnExcel(diccionario, pagina):
    columnas = "ABCDEFGHIJKLRSTUVWX"
    keys = list( diccionario.keys() )
    
    wb = openpyxl.load_workbook('Menta.xlsx')
    sheet = wb.create_sheet(pagina)
    
    for i in range(0,keys.__len__()):
        sheet[columnas[i] +'1'] = keys[i]
        fila = 2
        for v in list(diccionario.get(keys[i])):
            sheet[columnas[i] + fila.__str__()] = v
            fila = fila + 1
    
    wb.save('Menta.xlsx') #Fortato ods de OpenOffice da corrupto

def objetearEspectro(archivo):
    diccionario = {'key':[]}
    f = open(archivo + ".txt","r")
    for line in f.readlines():
        aux = line.rstrip().split(',')
        if all(map(lambda x: x.isalpha(),aux)):
            for variable in aux:
                diccionario[variable] = []
                diccionario['key'].append(variable)
        else:
            for variable,valor in zip(diccionario['key'],aux):
                diccionario[variable].append(float(valor))
    return diccionario
            
    
    
